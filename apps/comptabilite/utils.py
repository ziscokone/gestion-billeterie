"""
Utilitaires pour les exports comptables (Excel et PDF).
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


def format_montant(value):
    """Formate un montant avec séparateur de milliers."""
    if value is None:
        return "0"
    try:
        value = int(value)
        return f"{value:,}".replace(',', ' ')
    except (ValueError, TypeError):
        return "0"


def export_rapport_gare_excel(donnees, filtres):
    """
    Exporte le rapport par gare en format Excel.

    Args:
        donnees: Liste des lignes de données du rapport
        filtres: Dictionnaire contenant les filtres appliqués
            - gare_nom: Nom de la gare (ou "Toutes les gares")
            - ligne_nom: Nom de la ligne (ou "Toutes les lignes")
            - date_debut: Date de début
            - date_fin: Date de fin
            - total_charge: Total des dépenses
            - total_versement: Total du bénéfice net
            - types_depenses: Liste des types de dépenses présents
            - colonnes: Liste des colonnes à afficher

    Returns:
        HttpResponse avec le fichier Excel
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl n'est pas installé. Installez-le avec: pip install openpyxl")

    # Créer le workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport par gare"

    # Styles
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-tête du rapport
    row = 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = "RAPPORT PAR GARE"
    cell.font = Font(bold=True, size=16)
    cell.alignment = Alignment(horizontal='center')

    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    if filtres['date_debut'] == filtres['date_fin']:
        cell.value = f"Gare: {filtres['gare_nom']} | Date: {filtres['date_debut'].strftime('%d/%m/%Y')} | Ligne: {filtres['ligne_nom']}"
    else:
        cell.value = f"Gare: {filtres['gare_nom']} | Du {filtres['date_debut'].strftime('%d/%m/%Y')} au {filtres['date_fin'].strftime('%d/%m/%Y')} | Ligne: {filtres['ligne_nom']}"
    cell.alignment = Alignment(horizontal='center')

    row += 1
    ws.merge_cells(f'A{row}:H{row}')
    cell = ws[f'A{row}']
    cell.value = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    cell.font = Font(italic=True, size=9)
    cell.alignment = Alignment(horizontal='center')

    row += 2  # Ligne vide

    # En-têtes de colonnes
    col_index = 1
    for col_name in filtres['colonnes']:
        cell = ws.cell(row=row, column=col_index)
        cell.value = col_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        col_index += 1

    # Ajuster la largeur des colonnes
    column_widths = {
        'Date': 12,
        'Gare': 15,
        'Ligne': 25,
        'Num Départ': 12,
        'Nb Pass.': 10,
        'Recette Billets': 15,
        'Recette Bagages': 15,
        'Total Dépenses': 15,
        'Bénéfice Net': 15,
    }

    for idx, col_name in enumerate(filtres['colonnes'], start=1):
        width = column_widths.get(col_name, 14)
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

    # Données
    row += 1
    for ligne_data in donnees:
        col_index = 1
        for col_name in filtres['colonnes']:
            cell = ws.cell(row=row, column=col_index)
            value = ligne_data.get(col_name, 0)

            # Formatage selon le type de colonne
            if col_name in ['Recette Billets', 'Recette Bagages', 'Total Dépenses', 'Bénéfice Net'] or 'Carburant' in col_name or 'Frais' in col_name or 'Ration' in col_name or 'Réparation' in col_name or 'Divers' in col_name:
                # Colonnes monétaires
                cell.value = int(value) if value else 0
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            elif col_name == 'Nb Pass.':
                cell.value = int(value) if value else 0
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.value = value
                cell.alignment = Alignment(horizontal='left')

            cell.border = border
            col_index += 1
        row += 1

    # Ligne de total
    if donnees:
        col_index = 1
        for col_name in filtres['colonnes']:
            cell = ws.cell(row=row, column=col_index)
            cell.fill = total_fill
            cell.font = total_font
            cell.border = border

            if col_name in ['Date', 'Gare', 'Ligne', 'Num Départ']:
                if col_index == 1:
                    cell.value = "TOTAL"
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.value = ""
            elif col_name == 'Nb Pass.':
                cell.value = sum(int(d.get(col_name, 0)) for d in donnees)
                cell.alignment = Alignment(horizontal='center')
            elif col_name in ['Recette Billets', 'Recette Bagages', 'Total Dépenses', 'Bénéfice Net'] or any(x in col_name for x in ['Carburant', 'Frais', 'Ration', 'Réparation', 'Divers']):
                cell.value = sum(int(d.get(col_name, 0)) for d in donnees)
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = ""

            col_index += 1

        row += 2  # Lignes vides

        # Résumé final
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "CHARGE GARE (Total Dépenses):"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='right')

        cell = ws[f'D{row}']
        cell.value = int(filtres['total_charge'])
        cell.font = Font(bold=True, size=12)
        cell.number_format = '#,##0" FCFA"'

        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        cell = ws[f'A{row}']
        cell.value = "VERSEMENT (Bénéfice Net):"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='right')

        cell = ws[f'D{row}']
        cell.value = int(filtres['total_versement'])
        cell.font = Font(bold=True, size=12)
        cell.number_format = '#,##0" FCFA"'

    # Générer le fichier
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nom du fichier
    gare_slug = filtres['gare_nom'].lower().replace(' ', '-').replace('è', 'e').replace('é', 'e')
    date_str = filtres['date_debut'].strftime('%Y%m%d')
    if filtres['date_debut'] != filtres['date_fin']:
        date_str += f"_{filtres['date_fin'].strftime('%Y%m%d')}"
    filename = f"rapport_gare_{gare_slug}_{date_str}.xlsx"

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def export_rapport_gare_pdf(donnees, filtres):
    """
    Exporte le rapport par gare en format PDF avec un design professionnel.

    Args:
        donnees: Liste des lignes de données du rapport
        filtres: Dictionnaire contenant les filtres appliqués

    Returns:
        HttpResponse avec le fichier PDF
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab n'est pas installé. Installez-le avec: pip install reportlab")

    # Importer la compagnie
    try:
        from apps.compagnie.models import Compagnie
        compagnie = Compagnie.get_instance()
    except:
        compagnie = None

    # Créer le document
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=1*cm,
        bottomMargin=1*cm
    )

    # Conteneur pour les éléments
    elements = []
    styles = getSampleStyleSheet()

    # ========== EN-TÊTE ==========
    # Ligne de séparation supérieure
    header_line = Table([['']], colWidths=[27*cm])
    header_line.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2c3e50')),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(header_line)
    elements.append(Spacer(1, 0.2*cm))

    # Informations de la compagnie
    if compagnie:
        company_style = ParagraphStyle(
            'CompanyStyle',
            parent=styles['Normal'],
            fontSize=12,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=2
        )
        elements.append(Paragraph(f"<b>{compagnie.nom}</b>", company_style))

        if compagnie.adresse or compagnie.telephone:
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.HexColor('#7f8c8d'),
                spaceAfter=2
            )
            if compagnie.adresse:
                elements.append(Paragraph(compagnie.adresse, info_style))
            if compagnie.telephone:
                elements.append(Paragraph(f"Tél: {compagnie.telephone}", info_style))

    elements.append(Spacer(1, 0.3*cm))

    # ========== TITRE DU DOCUMENT ==========
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=8,
        alignment=1,  # Centre
        fontName='Helvetica-Bold'
    )
    elements.append(Paragraph("RAPPORT PAR GARE", title_style))

    # ========== INFORMATIONS DU RAPPORT ==========
    info_bg = Table([['']], colWidths=[27*cm], rowHeights=[0.8*cm])
    info_bg_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#ecf0f1')),
        ('ROUNDEDCORNERS', [3, 3, 3, 3]),
    ])
    info_bg.setStyle(info_bg_style)
    elements.append(info_bg)
    elements.append(Spacer(1, -0.7*cm))

    # Contenu des informations
    info_data = []
    if filtres['date_debut'] == filtres['date_fin']:
        date_text = f"Date: {filtres['date_debut'].strftime('%d/%m/%Y')}"
    else:
        date_text = f"Du {filtres['date_debut'].strftime('%d/%m/%Y')} au {filtres['date_fin'].strftime('%d/%m/%Y')}"

    info_data.append([
        f"<b>Gare:</b> {filtres['gare_nom']}",
        f"<b>{date_text}</b>",
        f"<b>Ligne:</b> {filtres['ligne_nom']}"
    ])

    info_table = Table(info_data, colWidths=[9*cm, 9*cm, 9*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2c3e50')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4*cm))

    # ========== TABLEAU PRINCIPAL ==========
    table_data = []

    # En-têtes
    headers = filtres['colonnes']
    table_data.append(headers)

    # Données
    for ligne_data in donnees:
        row = []
        for col_name in headers:
            value = ligne_data.get(col_name, 0)
            if col_name in ['Recette Billets', 'Recette Bagages', 'Total Dépenses', 'Bénéfice Net'] or any(x in col_name for x in ['Carburant', 'Frais', 'Ration', 'Réparation', 'Divers']):
                row.append(format_montant(value))
            else:
                row.append(str(value))
        table_data.append(row)

    # Ligne de total
    if donnees:
        total_row = []
        for col_name in headers:
            if col_name in ['Date', 'Gare', 'Ligne']:
                total_row.append("")
            elif col_name == 'Num Départ':
                total_row.append("TOTAL")
            elif col_name == 'Nb Pass.':
                total_row.append(str(sum(int(d.get(col_name, 0)) for d in donnees)))
            elif col_name in ['Recette Billets', 'Recette Bagages', 'Total Dépenses', 'Bénéfice Net'] or any(x in col_name for x in ['Carburant', 'Frais', 'Ration', 'Réparation', 'Divers']):
                total_row.append(format_montant(sum(int(d.get(col_name, 0)) for d in donnees)))
            else:
                total_row.append("")
        table_data.append(total_row)

    # Créer le tableau avec largeurs de colonnes dynamiques
    col_count = len(headers)
    col_width = 27*cm / col_count if col_count > 0 else 2*cm

    table = Table(table_data, colWidths=[col_width] * col_count)

    # Style du tableau
    table_style = TableStyle([
        # En-tête
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 5),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 5),

        # Données - lignes paires avec fond gris clair
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 7),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

        # Ligne de total
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 8),

        # Bordures
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#ddd')),
        ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#2c3e50')),

        # Padding
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ])

    # Alterner les couleurs des lignes
    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f8f9fa'))

    # Aligner les montants à droite
    for col_idx, col_name in enumerate(headers):
        if col_name in ['Recette Billets', 'Recette Bagages', 'Total Dépenses', 'Bénéfice Net', 'Nb Pass.'] or any(x in col_name for x in ['Carburant', 'Frais', 'Ration', 'Réparation', 'Divers']):
            table_style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'RIGHT')
        if col_name == 'Num Départ':
            table_style.add('ALIGN', (col_idx, 1), (col_idx, -1), 'CENTER')

    table.setStyle(table_style)
    elements.append(table)

    # ========== RÉSUMÉ FINAL ==========
    elements.append(Spacer(1, 0.5*cm))

    # Section avec fond coloré pour le résumé
    summary_data = [
        ['CHARGE GARE (Total Dépenses)', f"{format_montant(filtres['total_charge'])} FCFA"],
        ['', ''],  # Ligne de séparation
        ['VERSEMENT (Bénéfice Net)', f"{format_montant(filtres['total_versement'])} FCFA"]
    ]

    summary_table = Table(summary_data, colWidths=[13.5*cm, 13.5*cm])
    summary_style = TableStyle([
        # Styles généraux
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),

        # Charge Gare (rouge clair)
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#fadbd8')),
        ('TEXTCOLOR', (0, 0), (0, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (1, 0), (1, 0), colors.HexColor('#e74c3c')),
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
        ('ALIGN', (1, 0), (1, 0), 'CENTER'),

        # Ligne de séparation
        ('BACKGROUND', (0, 1), (-1, 1), colors.white),
        ('LINEABOVE', (0, 1), (-1, 1), 0, colors.white),
        ('LINEBELOW', (0, 1), (-1, 1), 0, colors.white),
        ('TOPPADDING', (0, 1), (-1, 1), 2),
        ('BOTTOMPADDING', (0, 1), (-1, 1), 2),

        # Versement (vert clair)
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#d5f4e6')),
        ('TEXTCOLOR', (0, 2), (0, 2), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor('#27ae60')),
        ('ALIGN', (0, 2), (0, 2), 'CENTER'),
        ('ALIGN', (1, 2), (1, 2), 'CENTER'),

        # Bordures
        ('BOX', (0, 0), (-1, 0), 1, colors.HexColor('#e74c3c')),
        ('BOX', (0, 2), (-1, 2), 1, colors.HexColor('#27ae60')),
    ])
    summary_table.setStyle(summary_style)
    elements.append(summary_table)

    # ========== PIED DE PAGE ==========
    elements.append(Spacer(1, 0.5*cm))

    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=7,
        textColor=colors.HexColor('#7f8c8d'),
        alignment=1
    )
    footer_text = f"Document généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    elements.append(Paragraph(footer_text, footer_style))

    # Générer le PDF
    doc.build(elements)
    output.seek(0)

    # Nom du fichier
    gare_slug = filtres['gare_nom'].lower().replace(' ', '-').replace('è', 'e').replace('é', 'e')
    date_str = filtres['date_debut'].strftime('%Y%m%d')
    if filtres['date_debut'] != filtres['date_fin']:
        date_str += f"_{filtres['date_fin'].strftime('%Y%m%d')}"
    filename = f"rapport_gare_{gare_slug}_{date_str}.pdf"

    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response
